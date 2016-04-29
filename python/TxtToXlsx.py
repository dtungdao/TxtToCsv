# import libraries
from openpyxl import Workbook # export excel file
import re
import os, sys


# global constants
INPUT_FILE_NAME = ""
OUTPUT_FILE_NAME = ""
BLOCK_INDICATOR = ("*************************************************************************************************")


# global variables
data_block = [] # hold text block
wb = Workbook()
ws = wb.active # grab the active worksheet


# get user input file (to be defined later)
if(len(sys.argv) < 3):
	print("USAGE: TxtToXlsx.py <txt_input_file_path> <xlsx_output_file_path>")
	sys.exit()
else:
	INPUT_FILE_NAME = sys.argv[1]
	OUTPUT_FILE_NAME = sys.argv[2]
	if(not os.path.isfile(INPUT_FILE_NAME)): 
		print("ERROR: the input file is not exist")
		sys.exit()


# get block data from input file
with open(INPUT_FILE_NAME, "rt") as f:
	txt = f.read()
	data_block = txt.split(BLOCK_INDICATOR)


# export block by block into XLSX file
row_idx = 1
for block in data_block:
	# create XLSX first row
	ws['A' + str(row_idx)] = "DF Study #"
	ws['B' + str(row_idx)] = "Time Stamp"
	ws['C' + str(row_idx)] = "Work FlowLevel"
	ws['D' + str(row_idx)] = "Data Records Pending"
	ws['E' + str(row_idx)] = "Data Records Lost"
	ws['F' + str(row_idx)] = "Data Records Incomplete"
	ws['G' + str(row_idx)] = "Data Records Page Final"
	ws['H' + str(row_idx)] = "Total"
	ws['I' + str(row_idx)] = "Queries Pending"
	ws['J' + str(row_idx)] = "Queries Unresolved"
	ws['K' + str(row_idx)] = "Queries Resolved"
	ws['L' + str(row_idx)] = "Queries Total"
	
	row_idx += 1 # new row

	lines = block.split("\n")
	
	df_study = 0;
	time_stamp = ""
	main_data = []
	awaiting_validation = 0
	being_validated = 0
	
	for line in lines:		
		if("#Database Status of Study " in line):
			df_study = int(re.findall(r'\b\d+\b', line)[0])
			ws['A' + str(row_idx)] = df_study
		elif("#Date		" in line):
			time_stamp = line[len("#Date		"):]
			ws['B' + str(row_idx)] = time_stamp
		elif(". Level " in line):
			main_data = re.findall(r'\b\d+\b', line[len(". Level "):])
			if(len(main_data) == 6):
				ws['C' + str(row_idx)] = main_data[0]
				ws['D' + str(row_idx)] = main_data[1]
				ws['E' + str(row_idx)] = main_data[2]
				ws['F' + str(row_idx)] = main_data[3]
				ws['G' + str(row_idx)] = main_data[4]
				ws['H' + str(row_idx)] = main_data[5]
			row_idx += 1 # new row
		elif("#Records awaiting validation" in line):
			awaiting_validation = int(re.findall(r'\b\d+\b', line)[0])
			ws['C' + str(row_idx)] = "Data records awaiting validation"
			ws['D' + str(row_idx)] = awaiting_validation
			row_idx += 1 # new row
		elif("#Records being validated" in line):
			being_validated = int(re.findall(r'\b\d+\b', line)[0])
			ws['C' + str(row_idx)] = "Data records being validated"
			ws['D' + str(row_idx)] = awaiting_validation
			row_idx += 2 # new row and a empty row
		else: pass
		

# Save the file
if(os.name != 'posix'): # window environment
	wb.save(os.path.dirname(os.path.realpath(__file__)) + "\\" + OUTPUT_FILE_NAME)
else: # unix environment
	wb.save(os.path.dirname(os.path.realpath(__file__)) + "/" + OUTPUT_FILE_NAME)
